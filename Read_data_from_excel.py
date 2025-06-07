import pandas as pd
from openpyxl import load_workbook

file_path= "files/QA_Engineer_Skills_Priority.xlsx"
df=pd.read_excel(file_path)
print(df)

wb = load_workbook(filename=file_path)
sheet = wb.active

for row in sheet.iter_rows(values_only=True):
    print(row)